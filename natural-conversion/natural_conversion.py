import json
import logging
import os
import sys
from math import floor
from pathlib import Path
from pathlib import PurePath

import boto3
import dask
import distributed
import numpy as np
import openpyxl
import parallel_functions
import psutil
import rasterio
import rioxarray
import xarray as xr
from dask.distributed import Client
from dask.distributed import LocalCluster
from dask.distributed import Lock
from dask.distributed import progress

TESTING = False
N_WORKERS = 14
THREADS_PER_WORKER = 6

DATA_PATH = Path("/data")

CROPLANDS_S3_BUCKET = "trends.earth-private"
CROPLANDS_S3_PREFIX = "cropland"
CCI_S3_BUCKET = "trends.earth-private"
OUT_S3_BUCKET = "trends.earth-private"
OUT_S3_PREFIX = "esa-cci/transitions"

INITIAL_YEAR = 2011
FINAL_YEAR = 2019
CROPLANDS_INITIAL_FILE = "Croplands_300m_2011.tif"
CROPLANDS_FINAL_FILE = "Croplands_300m_2019.tif"
CCI_TRANSITIONS_FILE = "ESACCI-LC-L4-LCCS-Map-300m-P1Y-Transitions_2011-2019.tif"
CCI_INITIAL_FILE = "ESACCI-LC-L4-LCCS-Map-300m-P1Y-2011-v2.0.7.tif"

formatter = "[%(levelname)s] %(asctime)s - %(message)s [%(funcName)s %(lineno)d]"
logging.basicConfig(level=logging.INFO, format=formatter)

logging.getLogger("botocore").setLevel(logging.WARNING)
logging.getLogger("s3transfer").setLevel(logging.WARNING)
logging.getLogger("boto3").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("distributed.worker").setLevel(logging.ERROR)

logger = logging.getLogger(__name__)


with open(PurePath("/data/aws_credentials.json"), "r") as f:
    aws_creds = json.load(f)


def put_to_s3(filename: Path, bucket: str, prefix: str):
    client = boto3.client("s3")
    key = f"{prefix}/{filename.name}"
    logger.info(f"Uploading {filename} to s3 at {key}")
    client.upload_file(str(filename), bucket, key)


def get_from_s3(bucket, prefix, filename, out_path):
    client = boto3.client("s3")
    logger.info(f"Downloading {filename} from s3 to {out_path}")
    client.download_file(bucket, f"{prefix}/{filename}", out_path)


def _log_file_size(out_file):
    file_size = os.stat(out_file).st_size
    logger.info(
        f"File size for %s is %s GB", out_file, round(file_size / (1024 ** 3), 2)
    )


def ds_to_cog(ds, client):
    ds.rio.write_crs("EPSG:4326", inplace=True)

    if TESTING:
        testing_string = "_TEST"
    else:
        testing_string = ""

    out_file = (
        DATA_PATH
        / f"natural-conversion_300m_{INITIAL_YEAR}-{FINAL_YEAR}{testing_string}.tif"
    )
    logger.info(f"Writing {out_file}...")
    ds.rio.to_raster(
        out_file,
        driver="Gtiff",
        compress="LZW",
        lock=Lock("rio-write", client=client),
    )
    _log_file_size(out_file)
    put_to_s3(out_file, OUT_S3_BUCKET, OUT_S3_PREFIX)


def ds_to_netcdf(ds):
    if TESTING:
        testing_string = "_TEST"
    else:
        testing_string = ""

    out_file = (
        DATA_PATH
        / f"natural-conversion_300m_{INITIAL_YEAR}-{FINAL_YEAR}{testing_string}.nc"
    )
    logger.info(f"Writing {out_file}...")
    encoding_dict = {key: {"zlib": True, "complevel": 6} for key in ds.data_vars.keys()}
    write_job = ds.to_netcdf(out_file, encoding=encoding_dict, compute=False)

    write_job = write_job.persist()
    progress(write_job)
    write_job.compute()
    _log_file_size(out_file)

    put_to_s3(out_file, OUT_S3_BUCKET, OUT_S3_PREFIX)


def ds_to_cogs(ds, client):
    ds.rio.write_crs("EPSG:4326", inplace=True)

    if TESTING:
        testing_string = "_TEST"
    else:
        testing_string = ""

    for name, array in ds.data_vars.items():
        out_file = (
            DATA_PATH
            / f"natural-conversion_300m_{INITIAL_YEAR}-{FINAL_YEAR}_{name}{testing_string}.tif"
        )
        array.rio.to_raster(
            out_file,
            driver="Gtiff",
            compress="LZW",
            lock=Lock("rio-write", client=client),
        )
        _log_file_size(out_file)
        put_to_s3(out_file, OUT_S3_BUCKET, OUT_S3_PREFIX)


def get_trans_codes(
    xl_file,
    initial_class_column,
    final_class_column,
    first_data_row,
    last_data_row,
):
    wb = openpyxl.load_workbook(xl_file)
    sheet = wb["Legend"]

    initial_class_codes = [
        val[0]
        for val in sheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=initial_class_column,
            max_col=initial_class_column,
            values_only=True,
        )
    ]

    final_class_codes = [
        val[0]
        for val in sheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=final_class_column,
            max_col=final_class_column,
            values_only=True,
        )
    ]

    return initial_class_codes, final_class_codes


def main():
    logger.info(
        "Using dask version %s, xarray version %s, rasterio version %s, "
        "rioxarray version %s, distributed version %s",
        dask.__version__,
        xr.__version__,
        rasterio.__version__,
        rioxarray.__version__,
        distributed.__version__,
    )

    # Download data
    DATA_PATH.mkdir(parents=True, exist_ok=True)

    crops_in_files = []
    for in_file in [CROPLANDS_INITIAL_FILE, CROPLANDS_FINAL_FILE]:
        local_crop_file_path = DATA_PATH / in_file
        crops_in_files.append(local_crop_file_path)

        if not local_crop_file_path.exists():
            get_from_s3(
                CROPLANDS_S3_BUCKET,
                CROPLANDS_S3_PREFIX,
                in_file,
                str(local_crop_file_path),
            )

    local_trans_file_path = DATA_PATH / CCI_TRANSITIONS_FILE
    if not local_trans_file_path.exists():
        get_from_s3(
            CCI_S3_BUCKET,
            "esa-cci/transitions",
            CCI_TRANSITIONS_FILE,
            str(local_trans_file_path),
        )

    local_initial_cover_file_path = DATA_PATH / CCI_INITIAL_FILE
    if not local_initial_cover_file_path.exists():
        get_from_s3(
            CCI_S3_BUCKET,
            "esa-cci",
            CCI_INITIAL_FILE,
            str(local_initial_cover_file_path),
        )

    logger.info("Loading data")

    with LocalCluster() as cluster, Client(cluster) as client:
        logger.info(f"cluster {cluster}")

        trans = rioxarray.open_rasterio(
            local_trans_file_path,
            chunks=dict(x=1024, y=1024),
            # lock=Lock("rio-read-trans", client=client),
        )
        # for trans band 1 is transition code, band 2 is meaning
        trans = trans.rename("trans").sel(band=2).drop("band")

        initial_cover = rioxarray.open_rasterio(
            local_initial_cover_file_path,
            chunks=dict(x=1024, y=1024),
            # lock=Lock("rio-read-initial-cover", client=client),
        )
        initial_cover = initial_cover.rename("lc_initial").sel(band=1).drop("band")

        crops_initial = rioxarray.open_rasterio(
            crops_in_files[0],
            chunks=dict(x=1024, y=1024),
            # lock=Lock("rio-read-crops", client=client),
        )
        crops_initial = crops_initial.rename("crops_initial").sel(band=1).drop("band")
        crops_final = rioxarray.open_rasterio(
            crops_in_files[-1],
            chunks=dict(x=1024, y=1024),
            # lock=Lock("rio-read-crops", client=client),
        )
        crops_final = crops_final.rename("crops_final").sel(band=1).drop("band")

        # Crop data for testing
        if TESTING:
            logger.warning("****** Cropping data for testing ******")
            trans = trans[22000:32000, 22000:32000]
            initial_cover = initial_cover[22000:32000, 22000:32000]
            crops_initial = crops_initial[22000:32000, 22000:32000]
            crops_final = crops_final[22000:32000, 22000:32000]

        in_data = xr.merge(
            [trans, initial_cover, crops_initial, crops_final],
            join="override",
            combine_attrs="drop",
        ).chunk(dict(x=512, y=512))

        trans_codes, trans_meanings = get_trans_codes(
            "ESA_CCI_Natural_Conversion_Coding_v2.xlsx",
            initial_class_column=1,
            final_class_column=3,
            first_data_row=3,
            last_data_row=40,
        )

        ###########################################################################
        # Compute transitions

        logger.info("Calculating natural conversion...")
        logger.info("in_data %s", in_data)

        # in_data = client.persist(in_data)

        # sys.exit()

        logger.info("Mapping compute_natural_conversion...")
        out = xr.map_blocks(
            parallel_functions.compute_natural_conversion,
            in_data,
            kwargs={
                "trans_codes": trans_codes,
                "trans_meanings": trans_meanings,
                "x_res": float((in_data.x[1] - in_data.x[0]).values),
                "y_res": float((in_data.y[0] - in_data.y[1]).values),
            },
        )

        ds_to_netcdf(out)

        # nat_conv = client.persist(nat_conv)
        # nat_conv = nat_conv.compute()


if __name__ == "__main__":
    # if TESTING:
    #     cluster = LocalCluster()
    # else:
    #     # total_memory = round(psutil.virtual_memory().total / (1024 ** 3), 2)
    #     # worker_memory = floor(total_memory / N_WORKERS)
    #     # logger.info(
    #     #     f"System has {total_memory} GB total memory. Using {N_WORKERS} workers, "
    #     #     f"with {THREADS_PER_WORKER} threads per worker, and {worker_memory} GB "
    #     #     "memory per worker."
    #     # )
    #     # cluster = LocalCluster(
    #     #     n_workers=N_WORKERS,
    #     #     threads_per_worker=THREADS_PER_WORKER,
    #     #     memory_limit=worker_memory,
    #     # )
    #     cluster = LocalCluster()
    # client = Client(cluster)

    main()
