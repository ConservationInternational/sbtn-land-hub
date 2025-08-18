import json
import logging
import os
import sys
from math import floor
from pathlib import Path
from pathlib import PurePath

import boto3
import dask
import distributed
import numpy as np
import openpyxl
import parallel_functions
import psutil
import rasterio
import rioxarray
import xarray as xr
from dask.distributed import Client
from dask.distributed import LocalCluster
from dask.distributed import Lock
from dask.distributed import progress

TESTING = False
N_WORKERS = 32
THREADS_PER_WORKER = 4

# TESTING = True
# N_WORKERS = 8
# THREADS_PER_WORKER = 4

DATA_PATH = Path("/data")

CCI_S3_BUCKET = "trends.earth-private"
OUT_S3_BUCKET = "trends.earth-private"
OUT_S3_PREFIX = "esa-cci/transitions"

INITIAL_YEAR = 2010
CCI_INITIAL_FILE = f"ESACCI-LC-L4-LCCS-Map-300m-P1Y-{INITIAL_YEAR}-v2.0.7.tif"

formatter = "[%(levelname)s] %(asctime)s - %(message)s [%(funcName)s %(lineno)d]"
logging.basicConfig(level=logging.INFO, format=formatter)

logging.getLogger("botocore").setLevel(logging.WARNING)
logging.getLogger("s3transfer").setLevel(logging.WARNING)
logging.getLogger("boto3").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("distributed.worker").setLevel(logging.ERROR)

logger = logging.getLogger(__name__)


with open(PurePath("/data/aws_credentials.json"), "r") as f:
    aws_creds = json.load(f)


def put_to_s3(filename: Path, bucket: str, prefix: str):
    client = boto3.client("s3")
    key = f"{prefix}/{filename.name}"
    logger.info(f"Uploading {filename} to s3 at {key}")
    client.upload_file(str(filename), bucket, key)


def get_from_s3(bucket, prefix, filename, out_path):
    client = boto3.client("s3")
    logger.info(f"Downloading {filename} from s3 to {out_path}")
    client.download_file(bucket, f"{prefix}/{filename}", out_path)


def _log_file_size(out_file):
    file_size = os.stat(out_file).st_size
    logger.info(
        f"File size for %s is %s GB", out_file, round(file_size / (1024 ** 3), 2)
    )


# def ds_to_cog(ds, client):
def ds_to_cog(ds):
    ds.rio.write_crs("EPSG:4326", inplace=True)

    if TESTING:
        testing_string = "_TEST"
    else:
        testing_string = ""

    out_file = (
        DATA_PATH / f"ESA-CCI-recoded_land_cover_{INITIAL_YEAR}{testing_string}.tif"
    )
    logger.info(f"Writing {out_file}...")
    ds.rio.to_raster(
        out_file,
        driver="Gtiff",
        compress="LZW",
        #    lock=Lock("rio-write", client=client),
    )
    _log_file_size(out_file)
    put_to_s3(out_file, OUT_S3_BUCKET, OUT_S3_PREFIX)


def get_recode(
    xl_file,
    initial_class_column,
    final_class_column,
    first_data_row,
    last_data_row,
):
    wb = openpyxl.load_workbook(xl_file)
    sheet = wb["Legend"]

    initial_class_codes = [
        val[0]
        for val in sheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=initial_class_column,
            max_col=initial_class_column,
            values_only=True,
        )
    ]

    final_class_codes = [
        val[0]
        for val in sheet.iter_rows(
            min_row=first_data_row,
            max_row=last_data_row,
            min_col=final_class_column,
            max_col=final_class_column,
            values_only=True,
        )
    ]

    return initial_class_codes, final_class_codes


def main():
    logger.info(
        "Using dask version %s, xarray version %s, rasterio version %s, "
        "rioxarray version %s, distributed version %s",
        dask.__version__,
        xr.__version__,
        rasterio.__version__,
        rioxarray.__version__,
        distributed.__version__,
    )

    # Download data
    DATA_PATH.mkdir(parents=True, exist_ok=True)

    local_initial_cover_file_path = DATA_PATH / CCI_INITIAL_FILE
    if not local_initial_cover_file_path.exists():
        get_from_s3(
            CCI_S3_BUCKET,
            "esa-cci",
            CCI_INITIAL_FILE,
            str(local_initial_cover_file_path),
        )

    logger.info("Loading data")

    with LocalCluster() as cluster, Client(cluster) as client:
        logger.info(f"cluster {cluster}")

        initial_cover = rioxarray.open_rasterio(
            local_initial_cover_file_path,
            chunks=dict(x=1024, y=1024),
            # lock=Lock("rio-read-initial-cover", client=client),
        )
        initial_cover = initial_cover.rename("lc_initial").sel(band=1).drop("band")

        # Crop data for testing
        if TESTING:
            logger.warning("****** Cropping data for testing ******")
            initial_cover = initial_cover[22000:32000, 22000:32000]

        initial_code, recode = get_recode(
            "ESA_CCI_Natural_Conversion_Coding_v2.xlsx",
            initial_class_column=1,
            final_class_column=3,
            first_data_row=3,
            last_data_row=40,
        )

        ###########################################################################
        # Compute transitions

        logger.info("Calculating natural conversion...")
        logger.info("initial_cover %s", initial_cover)

        logger.info("Mapping compute_natural_conversion...")
        out = xr.map_blocks(
            parallel_functions.recode_cover,
            initial_cover,
            kwargs={"initial_code": initial_code, "recode": recode},
        )

        # ds_to_cog(out, client)
        ds_to_cog(out)


if __name__ == "__main__":
    # if TESTING:
    #     cluster = LocalCluster()
    # else:
    #     # total_memory = round(psutil.virtual_memory().total / (1024 ** 3), 2)
    #     # worker_memory = floor(total_memory / N_WORKERS)
    #     # logger.info(
    #     #     f"System has {total_memory} GB total memory. Using {N_WORKERS} workers, "
    #     #     f"with {THREADS_PER_WORKER} threads per worker, and {worker_memory} GB "
    #     #     "memory per worker."
    #     # )
    #     # cluster = LocalCluster(
    #     #     n_workers=N_WORKERS,
    #     #     threads_per_worker=THREADS_PER_WORKER,
    #     #     memory_limit=worker_memory,
    #     # )
    #     cluster = LocalCluster()
    # client = Client(cluster)

    main()
